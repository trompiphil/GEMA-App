import streamlit as st
import pandas as pd
import gspread
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload
import datetime
import time
import os
import openpyxl
from openpyxl.cell.cell import MergedCell
from io import BytesIO

# --- KONFIGURATION ---
DB_NAME = "GEMA_Datenbank"

st.set_page_config(page_title="GEMA Manager", page_icon="xj", layout="centered")

# --- INITIALISIERUNG & RESET-LOGIK ---

def reset_draft_logic(keep_download=False):
    """Setzt das Formular zurück."""
    st.session_state.gig_draft = {
        "event_id": None, "datum": datetime.date.today(), "uhrzeit": datetime.time(19, 0),
        "ensemble": "Tutti", "location_selection": "Bitte wählen...", 
        "new_loc_data": {}
    }
    st.session_state.gig_song_selector = []
    
    if not keep_download:
        st.session_state.last_download = None
        st.session_state.uploaded_file_link = None

# State Variablen
if 'gig_draft' not in st.session_state: reset_draft_logic()
if 'gig_song_selector' not in st.session_state: st.session_state.gig_song_selector = []
if 'rep_edit_state' not in st.session_state: st.session_state.rep_edit_state = {"id": None, "titel": "", "dauer": "", "kn": "", "kv": "", "bn": "", "bv": "", "verlag": ""}
if 'page' not in st.session_state: st.session_state.page = "speichern"
if 'db_checked' not in st.session_state: st.session_state.db_checked = False
if 'last_download' not in st.session_state: st.session_state.last_download = None
if 'uploaded_file_link' not in st.session_state: st.session_state.uploaded_file_link = None
if 'trigger_reset' not in st.session_state: st.session_state.trigger_reset = False

# Reset Auslöser
if st.session_state.trigger_reset:
    reset_draft_logic(keep_download=True)
    st.session_state.trigger_reset = False
    st.rerun()

@st.cache_resource
def get_gspread_client():
    scopes = ["https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive"]
    s_info = st.secrets["gcp_service_account"]
    creds = Credentials.from_service_account_info(s_info, scopes=scopes)
    client = gspread.authorize(creds)
    drive_service = build('drive', 'v3', credentials=creds)
    return client, drive_service

try:
    client, drive_service = get_gspread_client()
    sh = client.open(DB_NAME)
except Exception as e:
    st.error(f"Verbindungsfehler: {e}"); st.stop()

# --- HELPER FUNKTIONEN (DRIVE) ---

def get_folder_id(folder_name, parent_id=None):
    """Sucht Ordner ID, optional in einem bestimmten Eltern-Ordner"""
    query = f"name = '{folder_name}' and mimeType = 'application/vnd.google-apps.folder' and trashed = false"
    if parent_id:
        query += f" and '{parent_id}' in parents"
    
    try:
        results = drive_service.files().list(q=query, fields="files(id)").execute()
        items = results.get('files', [])
        if items: return items[0]['id']
    except: pass
    return None

def list_files_in_templates():
    # Wir suchen erst den Hauptordner, um sicherzugehen
    root_id = get_folder_id("GEMA Bpol")
    if not root_id: return [], "Hauptordner 'GEMA Bpol' nicht gefunden. Bitte Ordner erstellen und mit Bot teilen!"
    
    # Dann Templates darin
    fid = get_folder_id("Templates", parent_id=root_id)
    if not fid: 
        # Fallback: Suche Templates global
        fid = get_folder_id("Templates")
        if not fid: return [], "Ordner 'Templates' nicht gefunden."

    query = f"'{fid}' in parents and trashed = false"
    results = drive_service.files().list(q=query, fields="files(id, name)").execute()
    return results.get('files', []), None

def download_specific_template(file_id, local_filename):
    try:
        content = drive_service.files().get_media(fileId=file_id).execute()
        with open(local_filename, "wb") as f:
            f.write(content)
        return True, None
    except Exception as e: return False, str(e)

# --- EXCEL HELPER ---

def safe_write(ws, row, col, value):
    """Schreibt sicher in eine Zelle, auch wenn sie verbunden (merged) ist."""
    try:
        cell = ws.cell(row=row, column=col)
        if isinstance(cell, MergedCell):
            for merged_range in ws.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    top_left = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                    top_left.value = value
                    break
        else:
            cell.value = value
    except Exception:
        pass # Ignoriere Schreibfehler in geschützte Bereiche

def process_and_upload_excel(template_file_id, datum, uhrzeit, ensemble, ort_data, songs_list, target_filename):
    # 1. Download Template lokal
    local_temp = "temp_template.xlsx"
    ok, err = download_specific_template(template_file_id, local_temp)
    if not ok: return None, None, f"Download Fehler: {err}"

    # 2. Excel bearbeiten
    try:
        wb = openpyxl.load_workbook(local_temp)
        ws = wb.active 
        
        # --- KOPFZEILEN 1-20 ---
        # User Wunsch: "Spalten A-Y, Zeilen 1-20 werden auf keinen Fall angerührt."
        # Daher schreibe ich hier KEINE Header-Daten rein, um das Layout nicht zu zerstören.
        
        # --- LISTE AB ZEILE 21 ---
        start_row = 21
        current_row = start_row
        
        # Leeren der alten Liste (nur sicherheitshalber bis Zeile 100)
        # Wir löschen nur Inhalte in den relevanten Spalten, um Formeln nicht zu zerstören
        cols_to_clear = [2, 5, 6, 7, 10, 16, 17] # B, E, F, G, J, P, Q
        for r in range(start_row, 100):
            for c in cols_to_clear:
                safe_write(ws, r, c, None)

        # Befüllen
        for song in songs_list:
            # Titel -> B (2)
            safe_write(ws, current_row, 2, song['Titel'])
            
            # Dauer -> E (5)
            safe_write(ws, current_row, 5, song['Dauer'])
            
            # Komponist Nachname -> F (6)
            safe_write(ws, current_row, 6, song['Komponist_Nachname'])
            
            # Komponist Vorname -> G (7)
            safe_write(ws, current_row, 7, song['Komponist_Vorname'])
            
            # Verlag -> J (10)
            safe_write(ws, current_row, 10, song['Verlag'])
            
            # Bearbeiter Nachname -> P (16)
            safe_write(ws, current_row, 16, song['Bearbeiter_Nachname'])
            
            # Bearbeiter Vorname -> Q (17)
            safe_write(ws, current_row, 17, song['Bearbeiter_Vorname'])
            
            current_row += 1
            
        wb.save(target_filename)
        
        # Bytes für direkten Download vorbereiten
        output_bytes = BytesIO()
        with open(target_filename, "rb") as f:
            output_bytes.write(f.read())
        output_bytes.seek(0)
        
    except Exception as e:
        return None, None, f"Excel Verarbeitungsfehler: {e}"

    # 3. Upload in die Cloud (GEMA Bpol/Output)
    web_link = "Lokal"
    try:
        # A) Hauptordner finden
        root_id = get_folder_id("GEMA Bpol")
        if root_id:
            # B) Output Ordner finden oder erstellen
            output_id = get_folder_id("Output", parent_id=root_id)
            if not output_id:
                # Versuchen zu erstellen (das darf der Bot, wenn er Schreibrechte auf GEMA Bpol hat)
                file_metadata = {
                    'name': 'Output',
                    'mimeType': 'application/vnd.google-apps.folder',
                    'parents': [root_id]
                }
                folder = drive_service.files().create(body=file_metadata, fields='id').execute()
                output_id = folder.get('id')
            
            if output_id:
                # C) Datei hochladen
                file_metadata = {'name': target_filename, 'parents': [output_id]}
                media = MediaFileUpload(target_filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                
                file = drive_service.files().create(body=file_metadata, media_body=media, fields='id, webViewLink').execute()
                web_link = file.get('webViewLink')
            else:
                st.warning("Konnte Output-Ordner nicht erstellen/finden.")
        else:
            st.warning("Ordner 'GEMA Bpol' nicht gefunden. Speichere nur lokal.")

    except Exception as e:
        st.warning(f"Upload fehlgeschlagen (Quota oder Rechte): {e}")

    # Aufräumen
    if os.path.exists(local_temp): os.remove(local_temp)
    if os.path.exists(target_filename): os.remove(target_filename)

    return output_bytes, web_link, None

# --- DB FUNKTIONEN ---
def check_and_fix_db():
    if st.session_state.db_checked: return
    try: ws_rep = sh.worksheet("Repertoire")
    except: ws_rep = sh.add_worksheet(title="Repertoire", rows=100, cols=15)
    rep_headers = ['ID', 'Titel', 'Komponist_Nachname', 'Komponist_Vorname', 'Bearbeiter_Nachname', 'Bearbeiter_Vorname', 'Dauer', 'Verlag', 'Werkeart', 'ISWC']
    if not ws_rep.row_values(1): ws_rep.update('A1:J1', [rep_headers])
    
    try: ws_ev = sh.worksheet("Events")
    except: ws_ev = sh.add_worksheet(title="Events", rows=100, cols=15)
    event_headers = ['Event_ID', 'Datum', 'Uhrzeit', 'Ensemble', 'Location_Name', 'Strasse', 'PLZ', 'Stadt', 'Setlist_Name', 'Songs_IDs', 'File_Link']
    curr = ws_ev.row_values(1)
    if not curr or 'File_Link' not in curr: ws_ev.clear(); ws_ev.update('A1:K1', [event_headers]) 
    
    try: ws_loc = sh.worksheet("Locations")
    except: ws_loc = sh.add_worksheet(title="Locations", rows=50, cols=5)
    loc_headers = ['ID', 'Name', 'Strasse', 'PLZ', 'Stadt']
    if not ws_loc.row_values(1): ws_loc.update('A1:E1', [loc_headers])
    st.session_state.db_checked = True

@st.cache_data(ttl=600)
def get_data_repertoire():
    ws = sh.worksheet("Repertoire")
    data = ws.get_all_records()
    df = pd.DataFrame(data)
    required = ['ID', 'Titel', 'Komponist_Nachname', 'Bearbeiter_Nachname', 'Verlag', 'Komponist_Vorname', 'Bearbeiter_Vorname', 'Dauer']
    for col in required:
        if col not in df.columns: df[col] = ""
    if not df.empty:
        df['ID'] = df['ID'].apply(lambda x: str(int(float(str(x).replace(',','.')))) if str(x).replace(',','.',1).replace('.','',1).isdigit() else str(x))
        df['Label'] = df.apply(lambda row: f"{row['Titel']} ({row['Komponist_Nachname']})" + (f" / Arr: {row['Bearbeiter_Nachname']}" if row['Bearbeiter_Nachname'] else ""), axis=1)
    return df

@st.cache_data(ttl=600)
def get_data_locations():
    ws = sh.worksheet("Locations")
    data = ws.get_all_records()
    df = pd.DataFrame(data)
    if df.empty: return pd.DataFrame(columns=['ID', 'Name', 'Strasse', 'PLZ', 'Stadt'])
    return df

@st.cache_data(ttl=600)
def get_data_events():
    ws = sh.worksheet("Events")
    data = ws.get_all_records()
    df = pd.DataFrame(data)
    if not df.empty and 'Datum' in df.columns:
        df['Datum_Obj'] = pd.to_datetime(df['Datum'], format="%d.%m.%Y", errors='coerce')
    return df

def clear_all_caches():
    get_data_repertoire.clear(); get_data_locations.clear(); get_data_events.clear()

def clean_id_list_from_string(raw_input):
    if not raw_input: return []
    parts = str(raw_input).split(',')
    clean = []
    for p in parts:
        s = p.strip()
        if not s: continue
        try: clean.append(str(int(float(s))))
        except: clean.append(s)
    return clean

def save_song_direct(mode, song_id, titel, kn, kv, bn, bv, dauer, verlag):
    ws = sh.worksheet("Repertoire")
    if mode == "Neu":
        col_ids = ws.col_values(1)[1:] 
        ids = [int(x) for x in col_ids if str(x).isdigit()]
        new_id = max(ids) + 1 if ids else 1
        row = [new_id, titel, kn, kv, bn, bv, dauer, verlag, "U-Musik", ""]
        ws.append_row(row)
        msg = f"'{titel}' angelegt!"
    else: 
        try:
            cell = ws.find(str(song_id), in_column=1)
            r = cell.row
            ws.update(f"B{r}:H{r}", [[titel, kn, kv, bn, bv, dauer, verlag]])
            msg = f"'{titel}' aktualisiert!"
        except: return False, "Fehler beim Update"
    clear_all_caches(); return True, msg

def save_location_direct(name, strasse, plz, stadt):
    ws = sh.worksheet("Locations")
    col_ids = ws.col_values(1)[1:]
    ids = [int(x) for x in col_ids if str(x).isdigit()]
    new_id = max(ids) + 1 if ids else 1
    ws.append_row([new_id, name, strasse, plz, stadt])
    clear_all_caches(); return True

def update_event_in_db(event_id, row_data):
    ws = sh.worksheet("Events")
    try:
        cell = ws.find(str(event_id), in_column=1)
        row_num = cell.row
        ws.update(f"A{row_num}:K{row_num}", [[event_id] + row_data])
        clear_all_caches(); return True
    except: return False

# --- NAVIGATION ---
def navigation_bar():
    st.markdown("---")
    c1, c2, c3, c4 = st.columns(4)
    if c1.button("💾 Speichern / Edit", use_container_width=True, type="primary" if st.session_state.page == "speichern" else "secondary"): st.session_state.page = "speichern"; st.rerun()
    if c2.button("🎵 Repertoire", use_container_width=True, type="primary" if st.session_state.page == "repertoire" else "secondary"): st.session_state.page = "repertoire"; st.rerun()
    if c3.button("📍 Orte", use_container_width=True, type="primary" if st.session_state.page == "orte" else "secondary"): st.session_state.page = "orte"; st.rerun()
    if c4.button("📂 Archiv", use_container_width=True, type="primary" if st.session_state.page == "archiv" else "secondary"): st.session_state.page = "archiv"; st.rerun()
    st.markdown("---")

# --- MAIN ---
check_and_fix_db()
st.title("Orchester Manager 🎻")
navigation_bar()

# === SEITE 1: SPEICHERN ===
if st.session_state.page == "speichern":
    df_loc = get_data_locations()
    df_rep = get_data_repertoire()
    df_events = get_data_events()
    
    # 1. DOWNLOAD & CLOUD STATUS
    if st.session_state.last_download:
        d_name, d_bytes = st.session_state.last_download
        cloud_link = st.session_state.uploaded_file_link
        
        st.success("🎉 Verarbeitung abgeschlossen!")
        c_dl, c_cloud = st.columns(2)
        
        c_dl.download_button(
            label=f"📥 {d_name} herunterladen",
            data=d_bytes,
            file_name=d_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True
        )
        
        if cloud_link and cloud_link != "Lokal":
            c_cloud.link_button("☁️ In Google Drive öffnen", cloud_link, use_container_width=True)
        else:
            c_cloud.info("Nur lokal gespeichert (kein Cloud-Upload).")
            
        st.divider()

    # 2. Editier-Auswahl
    if st.session_state.gig_draft["event_id"] is None:
        with st.expander("🛠 Bereits gespeicherten Auftritt bearbeiten", expanded=False):
            if not df_events.empty:
                df_events['Label'] = df_events.apply(lambda x: f"{x['Datum']} - {x['Location_Name']} ({x['Ensemble']})", axis=1)
                df_events = df_events.sort_values(by='Datum_Obj', ascending=False)
                edit_sel = st.selectbox("Wähle einen Auftritt:", ["Bitte wählen..."] + df_events['Label'].tolist())
                if st.button("Laden") and edit_sel != "Bitte wählen...":
                    row = df_events[df_events['Label'] == edit_sel].iloc[0]
                    st.session_state.gig_draft["event_id"] = row['Event_ID']
                    st.session_state.gig_draft["datum"] = datetime.datetime.strptime(row['Datum'], "%d.%m.%Y").date()
                    try: st.session_state.gig_draft["uhrzeit"] = datetime.datetime.strptime(row['Uhrzeit'], "%H:%M").time()
                    except: st.session_state.gig_draft["uhrzeit"] = datetime.time(19, 0)
                    st.session_state.gig_draft["ensemble"] = row['Ensemble']
                    st.session_state.gig_draft["location_selection"] = row['Location_Name']
                    saved_ids = clean_id_list_from_string(row['Songs_IDs'])
                    restored_labels = []
                    if not df_rep.empty:
                        rows = df_rep[df_rep['ID'].isin(saved_ids)]
                        restored_labels = rows['Label'].tolist()
                    st.session_state.gig_song_selector = restored_labels
                    st.session_state.last_download = None 
                    st.toast("Geladen!", icon="✏️"); time.sleep(0.5); st.rerun()

    if st.session_state.gig_draft["event_id"]:
        cb, ch = st.columns([1,3])
        # Zurück löscht auch den Download Button
        if cb.button("⬅️ Zurück"): 
            st.session_state.trigger_reset = True
            st.rerun()
        ch.header(f"✏️ Bearbeiten (ID: {st.session_state.gig_draft['event_id']})")
    else: st.header("📝 Neuen Auftritt erfassen")

    c_date, c_time = st.columns(2)
    st.session_state.gig_draft["datum"] = c_date.date_input("Datum", value=st.session_state.gig_draft["datum"])
    st.session_state.gig_draft["uhrzeit"] = c_time.time_input("Uhrzeit", value=st.session_state.gig_draft["uhrzeit"])
    ens_opts = ["Tutti", "BQ", "Quartett", "Duo"]
    st.session_state.gig_draft["ensemble"] = st.selectbox("Ensemble", ens_opts, index=ens_opts.index(st.session_state.gig_draft["ensemble"]))

    st.write("📍 **Spielort**")
    loc_opts = ["Bitte wählen..."] + df_loc['Name'].tolist() + ["➕ Neuer Ort..."]
    try: sel_idx = loc_opts.index(st.session_state.gig_draft["location_selection"])
    except: sel_idx = 0
    sel_loc = st.selectbox("Ort:", options=loc_opts, index=sel_idx)
    st.session_state.gig_draft["location_selection"] = sel_loc

    final_loc = {}
    if sel_loc == "➕ Neuer Ort...":
        with st.form("new_loc_form"):
            n = st.text_input("Name*"); s = st.text_input("Str."); p = st.text_input("PLZ"); ci = st.text_input("Stadt*")
            if st.form_submit_button("Bestätigen"):
                if n and ci:
                    save_location_direct(n, s, p, ci)
                    st.session_state.gig_draft["location_selection"] = n # Auto select new location
                    st.toast(f"Ort '{n}' gespeichert!", icon="✅")
                    time.sleep(1)
                    st.rerun()
                else:
                    st.error("Name und Stadt sind Pflicht!")
    elif sel_loc != "Bitte wählen...":
        final_loc = df_loc[df_loc['Name'] == sel_loc].iloc[0].to_dict()

    st.markdown("---")
    st.write("🎵 **Programm**")
    with st.expander("➕ Titel fehlt? Hier sofort anlegen"):
        with st.form("quick_add"):
            c1, c2 = st.columns([3,1]); t=c1.text_input("Titel"); d=c2.text_input("Dauer", "03:00")
            c3, c4 = st.columns(2); kn=c3.text_input("Komp NN"); kv=c4.text_input("Komp VN")
            c5, c6 = st.columns(2); bn=c5.text_input("Bearb NN"); bv=c6.text_input("Bearb VN")
            ver = st.text_input("Verlag")
            if st.form_submit_button("Speichern"):
                if t and kn: save_song_direct("Neu", None, t, kn, kv, bn, bv, d, ver); st.rerun()

    if not df_rep.empty:
        selection = st.multiselect("Suche:", options=df_rep['Label'].tolist(), key="gig_song_selector")
        
        st.markdown("---")
        files_found, err_msg = list_files_in_templates()
        if not files_found: st.error("⚠️ 'Templates' Ordner ist leer!"); selected_template_id = None
        else:
            file_options = {f['name']: f['id'] for f in files_found}
            default_idx = 0
            for i, fname in enumerate(file_options.keys()):
                if "setlist" in fname.lower(): default_idx = i; break
            sel_temp_name = st.selectbox("Vorlage:", list(file_options.keys()), index=default_idx)
            selected_template_id = file_options[sel_temp_name]

        btn_txt = "🔄 Aktualisieren & Excel erstellen" if st.session_state.gig_draft["event_id"] else "✅ Final speichern & Excel erstellen"
        
        if st.button(btn_txt, type="primary", use_container_width=True):
            if not final_loc.get("Name") or not selection or not selected_template_id:
                st.error("Daten fehlen! (Ort, Programm oder Template)")
            else:
                # 1. Ort ggf. speichern (Doppelt hält besser falls User nicht Bestätigen klickte)
                if sel_loc == "➕ Neuer Ort...":
                    # Hier können wir nicht viel tun außer Fehler werfen, da User oben bestätigen muss
                    st.error("Bitte erst den neuen Ort bestätigen!")
                else:
                    datum_str = st.session_state.gig_draft["datum"].strftime("%d.%m.%Y")
                    time_str = st.session_state.gig_draft["uhrzeit"].strftime("%H:%M")
                    dateiname = f"{st.session_state.gig_draft['ensemble']}{datum_str}{final_loc['Stadt']}Setlist.xlsx"
                    
                    selected_songs_data = []
                    song_ids = []
                    for label in selection:
                        row = df_rep[df_rep['Label'] == label].iloc[0]
                        song_ids.append(str(row['ID']))
                        selected_songs_data.append(row.to_dict())
                    
                    with st.spinner("Erstelle Excel-Datei und lade hoch..."):
                        excel_bytes, web_link, err = process_and_upload_excel(
                            selected_template_id, 
                            datum_str, 
                            time_str, 
                            st.session_state.gig_draft['ensemble'], 
                            final_loc, 
                            selected_songs_data, 
                            dateiname
                        )
                        
                        if err:
                            st.error(f"⚠️ {err}")
                        else:
                            row_data = [
                                datum_str, time_str, st.session_state.gig_draft["ensemble"],
                                final_loc["Name"], final_loc["Strasse"], str(final_loc["PLZ"]), final_loc["Stadt"],
                                dateiname, ",".join(song_ids), str(web_link)
                            ]
                            
                            if st.session_state.gig_draft["event_id"]: update_event_in_db(st.session_state.gig_draft["event_id"], row_data)
                            else:
                                ws_ev = sh.worksheet("Events"); col_ids = ws_ev.col_values(1)[1:]
                                e_ids = [int(x) for x in col_ids if str(x).isdigit()]
                                new_ev_id = max(e_ids) + 1 if e_ids else 1
                                ws_ev.append_row([new_ev_id] + row_data)
                                clear_all_caches()
                            
                            st.session_state.last_download = (dateiname, excel_bytes.getvalue())
                            st.session_state.uploaded_file_link = web_link
                            st.session_state.trigger_reset = True
                            st.rerun()
    else: st.warning("Repertoire leer.")

# === SEITE 2,3,4 (Gleich) ===
elif st.session_state.page == "repertoire":
    st.subheader("Repertoire verwalten")
    mode = st.radio("Modus:", ["Neu", "Bearbeiten"], horizontal=True)
    df_rep = get_data_repertoire()
    if mode == "Bearbeiten" and not df_rep.empty:
        sel = st.selectbox("Suchen:", df_rep['Label'].tolist(), index=None)
        if sel:
            r = df_rep[df_rep['Label'] == sel].iloc[0]
            if st.session_state.rep_edit_state["id"] != r['ID']: st.session_state.rep_edit_state = {"id": r['ID'], "titel": r['Titel'], "dauer": str(r['Dauer']), "kn": r['Komponist_Nachname'], "kv": r['Komponist_Vorname'], "bn": r['Bearbeiter_Nachname'], "bv": r['Bearbeiter_Vorname'], "verlag": r['Verlag']}
    elif mode == "Neu" and st.session_state.rep_edit_state["id"]: st.session_state.rep_edit_state = {"id": None, "titel": "", "dauer": "03:00", "kn": "", "kv": "", "bn": "", "bv": "", "verlag": ""}
    with st.form("rep"):
        s = st.session_state.rep_edit_state
        c1,c2=st.columns([3,1]); t=c1.text_input("Titel", s["titel"]); d=c2.text_input("Dauer", s["dauer"])
        c3,c4=st.columns(2); kn=c3.text_input("Komp NN", s["kn"]); kv=c4.text_input("Komp VN", s["kv"])
        c5,c6=st.columns(2); bn=c5.text_input("Bearb NN", s["bn"]); bv=c6.text_input("Bearb VN", s["bv"])
        v=st.text_input("Verlag", s["verlag"])
        if st.form_submit_button("Speichern"):
            save_song_direct("Edit" if s["id"] else "Neu", s["id"], t, kn, kv, bn, bv, d, v); st.rerun()
    st.dataframe(df_rep, use_container_width=True)

elif st.session_state.page == "orte":
    st.subheader("Locations"); df_loc = get_data_locations()
    with st.form("loc"):
        n=st.text_input("Name"); ci=st.text_input("Stadt")
        if st.form_submit_button("Speichern"): save_location_direct(n, "", "", ci); st.rerun()
    st.dataframe(df_loc, use_container_width=True)

elif st.session_state.page == "archiv":
    st.subheader("📂 Setlist Archiv")
    df_events = get_data_events()
    if not df_events.empty:
        df_events = df_events.sort_values(by='Datum_Obj', ascending=False)
        for year in df_events['Datum_Obj'].dt.year.unique():
            st.markdown(f"### {year}")
            df_y = df_events[df_events['Datum_Obj'].dt.year == year]
            for m in df_y['Datum_Obj'].dt.month.unique():
                m_name = datetime.date(2000, int(m), 1).strftime('%B')
                with st.expander(f"{m_name} ({len(df_y[df_y['Datum_Obj'].dt.month == m])})"):
                    for _, row in df_y[df_y['Datum_Obj'].dt.month == m].iterrows():
                        c1, c2 = st.columns([3, 1])
                        c1.write(f"**{row['Datum']}** | {row['Location_Name']} ({row['Ensemble']})"); c1.caption(f"Datei: {row['Setlist_Name']}")
                        
                        if 'File_Link' in row and row['File_Link'] and row['File_Link'] != "Lokal":
                            c2.link_button("☁️ Drive", row['File_Link'])
                        else:
                            c2.caption("Nur Lokal")
                        st.divider()
